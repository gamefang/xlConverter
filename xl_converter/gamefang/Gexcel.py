# -*- coding: utf-8 -*-
# excel處理模塊
# 依賴
# pip install openpyxl

#region import
# 仅导入类型
from __future__ import annotations
from typing import TYPE_CHECKING
# std
import warnings
# site
from openpyxl import open as xl_open
if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.worksheet import Worksheet
#endregion

# 屏蔽openpyxl對數據驗證的報錯
warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")

class gExcel:
    '''
    excel處理模塊
    '''
    @staticmethod
    def get_workbook(fn: str, read_only: bool = True, data_only: bool = True) -> Workbook:
        '''
        獲取一個excel工作表
        '''
        return xl_open(fn, read_only = read_only, data_only = data_only)

    @staticmethod
    def get_named_cells(workbook: Workbook, defined_name: str) -> list:
        '''
        獲取工作表級別的命名區域
        '''
        named_range: DefinedName = workbook.defined_names[defined_name]
        dests = named_range.destinations
        cells = []
        for title, coord in dests:
            ws = workbook[title]
            cells.append(ws[coord])
        return cells[0] # 会多嵌套一层，自动解开

    @staticmethod
    def get_sheet_range(sheet: Worksheet, tag: str = '#') -> list[int]:
        '''
        檢測標記範圍內的sheet區域，返回列表[row_start, row_end, col_start, col_end]

        :param sheet: openpyxl的sheet
        :param tag: 標記符號，位於區域的右上角和左下角（不包含）
        '''
        row_start = row_end = col_start = col_end = 0
        for row_index, row_values in enumerate(sheet.iter_rows(values_only=True), start=1):
            try:
                col_index = list(row_values).index(tag) + 1
            except ValueError:
                continue
            if row_start == 0:
                row_start = row_index
                col_end = col_index - 1
            else:
                row_end = row_index - 1
                col_start = col_index
                break
        return [row_start, row_end, col_start, col_end]

    @staticmethod
    def get_sheet_range_value_list(sheet: Worksheet, tag: str = '#') -> list:
        '''
        獲取標記sheet區域內所有單元格的值，返回二維列表

        :param sheet: openpyxl的sheet
        :param tag: 標記符號，位於區域的右上角和左下角（不包含）
        '''
        result = []
        row_start, row_end, col_start, col_end = __class__.get_sheet_range(sheet, tag)
        for row in sheet.iter_rows(
            min_row = row_start, max_row = row_end,
            min_col = col_start, max_col = col_end,
            values_only = True, # 僅加載數值
        ):
            result.append(list(row))
        return result