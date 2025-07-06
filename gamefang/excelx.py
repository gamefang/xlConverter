# -*- coding: utf-8 -*-
# excel處理模塊

# 依賴
# pip install openpyxl

import openpyxl
import warnings
# 屏蔽openpyxl對數據驗證的報錯
warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")

def get_workbook(fn, read_only = True, data_only = True):
    '''
    獲取一個excel工作表
    '''
    return openpyxl.open(fn, read_only = read_only, data_only = data_only)

def get_named_cells(workbook, defined_name):
    '''
    獲取工作表級別的命名區域
    '''
    named_range = workbook.defined_names[defined_name]
    dests = named_range.destinations
    cells = []
    for title, coord in dests:
        ws = workbook[title]
        cells.append(ws[coord])
    return cells[0] # 会多嵌套一层，自动解开

def get_sheet_range(sheet, tag = '#'):
    '''
    檢測標記範圍內的sheet區域，返回列表[row_start, row_end, col_start, col_end]

    :param sheet: openpyxl的sheet
    :param tag: 標記符號，位於區域的右上角和左下角（不包含）
    '''
    max_row = sheet.max_row
    max_column = sheet.max_column
    row_start = row_end = col_start = col_end = 0
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value == tag:
                if row_start == 0:
                    row_start = row
                    col_end = col - 1
                else:
                    row_end = row - 1
                    col_start = col
            if row_start != 0 and row_end != 0:
                break
    return [row_start, row_end, col_start, col_end]

def get_sheet_range_value_list(sheet, tag = '#'):
    '''
    獲取標記sheet區域內所有單元格的值，返回二維列表

    :param sheet: openpyxl的sheet
    :param tag: 標記符號，位於區域的右上角和左下角（不包含）
    '''
    result = []
    row_start, row_end, col_start, col_end = get_sheet_range(sheet, tag)
    for row in sheet.iter_rows(
        min_row = row_start, max_row = row_end,
        min_col = col_start, max_col = col_end,
        values_only = True, # 僅加載數值
    ):
        result.append(list(row))
    return result